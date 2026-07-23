import openpyxl
import datetime
import os
import subprocess

def clean_value(val):
    if val is None:
        return "NULL"
    
    if isinstance(val, str):
        val_escaped = val.replace("'", "''").strip()
        if not val_escaped or val_escaped.lower() == 'none':
            return "NULL"
        return f"N'{val_escaped}'"
    
    if isinstance(val, datetime.datetime):
        return f"'{val.strftime('%Y-%m-%d')}'"
    
    return f"N'{str(val)}'"

def main():
    excel_path = r'c:\Users\CCL-04\ai_project\offshore roaster\April - July schedule.xlsx'
    sql_temp_path = r'c:\Users\CCL-04\ai_project\offshore roaster\insert_roster.sql'
    
    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found at {excel_path}")
        return
        
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['April 13  July 13 schedule']
    
    headers = [cell.value for cell in ws[1]]
    table_name = "CCL Offshore Team Shift Roaster"
    
    sql_lines = []
    sql_lines.append("USE OffshoreShiftRosterDB;")
    sql_lines.append("GO\n")
    
    # Drop and recreate the table to keep it fully synchronized
    sql_lines.append(f"IF OBJECT_ID('{table_name}', 'U') IS NOT NULL")
    sql_lines.append(f"    DROP TABLE [{table_name}];")
    sql_lines.append("GO\n")
    
    create_stmt = (
        f"CREATE TABLE [{table_name}] (\n"
        f"    [{headers[0]}] DATE PRIMARY KEY,\n"
        f"    [{headers[1]}] VARCHAR(50) NOT NULL,\n"
        f"    [{headers[2]}] NVARCHAR(500) NULL,\n"
        f"    [{headers[3]}] NVARCHAR(500) NULL,\n"
        f"    [{headers[4]}] NVARCHAR(500) NULL,\n"
        f"    [{headers[5]}] NVARCHAR(500) NULL,\n"
        f"    [{headers[6]}] NVARCHAR(1000) NULL\n"
        f");\n"
        f"GO\n"
    )
    sql_lines.append(create_stmt)
    
    dates_seen = set()
    rows_inserted = 0
    
    for r_idx in range(2, ws.max_row + 1):
        row = [cell.value for cell in ws[r_idx]]
        date_val = row[0]
        
        # Only calendar schedule rows start with a datetime object
        if isinstance(date_val, datetime.datetime):
            formatted_date = date_val.strftime('%Y-%m-%d')
            if formatted_date in dates_seen:
                continue
            dates_seen.add(formatted_date)
            
            day_val = row[1] if row[1] else date_val.strftime('%A')
            day_str = f"'{day_val}'"
            col3_str = clean_value(row[2])
            col4_str = clean_value(row[3])
            col5_str = clean_value(row[4])
            col6_str = clean_value(row[5])
            col7_str = clean_value(row[6])
            
            insert_stmt = (
                f"INSERT INTO [{table_name}] "
                f"([{headers[0]}], [{headers[1]}], [{headers[2]}], [{headers[3]}], [{headers[4]}], [{headers[5]}], [{headers[6]}]) "
                f"VALUES ('{formatted_date}', {day_str}, {col3_str}, {col4_str}, {col5_str}, {col6_str}, {col7_str});"
            )
            sql_lines.append(insert_stmt)
            rows_inserted += 1
            
    sql_lines.append("\nGO")
    
    with open(sql_temp_path, 'w', encoding='utf-8') as f:
        f.write("\n".join(sql_lines))
        
    print(f"Prepared SQL script with {rows_inserted} rows.")
    
    # Run sqlcmd to execute the script
    print("Executing sqlcmd...")
    try:
        result = subprocess.run(
            ['sqlcmd', '-S', '.\\SQLEXPRESS', '-E', '-C', '-i', sql_temp_path],
            capture_output=True,
            text=True,
            check=True
        )
        print("SQL Server execution output:")
        print(result.stdout)
        print("Synchronization completed successfully!")
    except subprocess.CalledProcessError as e:
        print("Error executing SQL script:")
        print(e.stderr)
        print(e.stdout)

if __name__ == '__main__':
    main()
