import openpyxl
import datetime
import os

def clean_value(val):
    if val is None:
        return "NULL"
    
    # Clean up value if string
    if isinstance(val, str):
        # Escape single quotes for SQL
        val_escaped = val.replace("'", "''").strip()
        if not val_escaped or val_escaped.lower() == 'none':
            return "NULL"
        return f"N'{val_escaped}'"
    
    if isinstance(val, datetime.datetime):
        return f"'{val.strftime('%Y-%m-%d')}'"
    
    return f"N'{str(val)}'"

def main():
    excel_path = r'c:\Users\CCL-04\ai_project\April - July schedule.xlsx'
    sql_output_path = r'c:\Users\CCL-04\ai_project\insert_roster.sql'
    
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['April 13  July 13 schedule']
    
    # Extract headers exactly as they are in row 1
    headers = [cell.value for cell in ws[1]]
    
    # Table name requested: CCL Offshore Team Shift Roaster
    table_name = "CCL Offshore Team Shift Roaster"
    
    sql_lines = []
    sql_lines.append("USE OffshoreShiftRosterDB;")
    sql_lines.append("GO\n")
    
    # Drop table if exists
    sql_lines.append(f"IF OBJECT_ID('{table_name}', 'U') IS NOT NULL")
    sql_lines.append(f"    DROP TABLE [{table_name}];")
    sql_lines.append("GO\n")
    
    # Create Table statement with exact column names from Excel
    create_stmt = (
        f"CREATE TABLE [{table_name}] (\n"
        f"    [{headers[0]}] DATE PRIMARY KEY,\n"
        f"    [{headers[1]}] VARCHAR(50) NOT NULL,\n"
        f"    [{headers[2]}] NVARCHAR(500) NULL,\n"
        f"    [{headers[3]}] NVARCHAR(500) NULL,\n"
        f"    [{headers[4]}] NVARCHAR(500) NULL,\n"
        f"    [{headers[5]}] NVARCHAR(500) NULL,\n"
        f"    [{headers[6]}] NVARCHAR(1000) NULL\n"
        f");\n"
        f"GO\n"
    )
    sql_lines.append(create_stmt)
    
    # Insert statements
    dates_seen = set()
    rows_inserted = 0
    
    for r_idx in range(2, ws.max_row + 1):
        row = [cell.value for cell in ws[r_idx]]
        date_val = row[0]
        
        # We only keep rows where the first cell is an actual datetime object
        if isinstance(date_val, datetime.datetime):
            formatted_date = date_val.strftime('%Y-%m-%d')
            if formatted_date in dates_seen:
                print(f"Warning: Duplicate date found at row {r_idx}: {formatted_date}")
                continue
            dates_seen.add(formatted_date)
            
            # Format other columns
            day_val = row[1]
            if not day_val:
                day_val = date_val.strftime('%A')
            
            day_str = f"'{day_val}'"
            col3_str = clean_value(row[2])
            col4_str = clean_value(row[3])
            col5_str = clean_value(row[4])
            col6_str = clean_value(row[5])
            col7_str = clean_value(row[6])
            
            insert_stmt = (
                f"INSERT INTO [{table_name}] "
                f"([{headers[0]}], [{headers[1]}], [{headers[2]}], [{headers[3]}], [{headers[4]}], [{headers[5]}], [{headers[6]}]) "
                f"VALUES ('{formatted_date}', {day_str}, {col3_str}, {col4_str}, {col5_str}, {col6_str}, {col7_str});"
            )
            sql_lines.append(insert_stmt)
            rows_inserted += 1
            
    sql_lines.append("\nGO")
    
    with open(sql_output_path, 'w', encoding='utf-8') as f:
        f.write("\n".join(sql_lines))
        
    print(f"SQL file written to: {sql_output_path}")
    print(f"Total rows prepared for insertion: {rows_inserted}")

if __name__ == '__main__':
    main()
